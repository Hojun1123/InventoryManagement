import openpyxl as op


def readfiletolist(file, sheet):
    rb = op.load_workbook("./DB/"+file+".xlsx")
    ws = rb[sheet]
    # print(ws['C4'].value)
    datalist = []
    for row in ws.rows:
        datalist.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])

    print(datalist)
    return datalist


def writelisttofile(data, path, sheet_name):
    wb = op.load_workbook(path)
    _ = wb.active
    sheet = wb[sheet_name]
    sheet.append(data)
    wb.save(path)
    wb.close()


def handler_test():
    readfiletolist('../DB/data.xlsx', 'Sheet1')
    writelisttofile([1, 2, 3, 4, 5], '../DB/data.xlsx', 'Sheet1')
    readfiletolist('../DB/data.xlsx', 'Sheet1')