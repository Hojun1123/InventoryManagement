# db control
import threading
from datetime import datetime
import openpyxl as op
from collections import defaultdict
import pandas as pd
import math
import numpy as np
import pandas as pd


def read_sheet(file, sheet):
    return pd.read_excel("./DB/" + file + ".xlsx", sheet, dtype=str)


def open_sheet(file, sheet):
    rb = op.load_workbook("./DB/" + file + ".xlsx")
    return rb[sheet]


def open_file(file):
    return op.load_workbook("./DB/" + file + ".xlsx")


def get_path(file):
    return "./DB/" + file + ".xlsx"


#마지막 동기화 시간 반환
def last_sync_time():
    try:
        rs = open_sheet("barcode", "rawBarcode")
    except:
        print("can't open barcode.xlsx")
        return -1
    return [i.value for i in list(rs.rows)[-1][1:3]]


def append_raw_barcodes(data):
    try:
        wb1 = open_file("barcode")
    except:
        print("can't open barcode.xlsx")
        return -1
    w1s1 = wb1["rawBarcode"]
    c = 0
    sl = len(list(w1s1.rows))
    tm = datetime.datetime.now()
    date = tm.strftime("%Y%m%d")
    time = (tm.strftime("%X"))[0:2] + (tm.strftime("%X"))[3:5] + (tm.strftime("%X"))[6:8]
    for i in data:
        c += 1
        gi = sl + c
        # gk, location
        for j in i:
            # 엔진시리얼번호, 바코드, 날짜, 시간, 그룹ID
            w1s1.append([j, date, time, str(gi)])
    wb1.save(get_path("barcode"))
    wb1.close()


def get_data_to_add(day, time):
    try:
        wb = open_sheet("barcode", "rawBarcode")
    except:
        print("can't open barcode.xlsx")
        return -1
    data = []
    dlist = list(wb.rows)
    for i in range(1, len(dlist)):
        if int(dlist[i][1].value) > int(day):
            #barcode, date, gid
            data.append([dlist[i][0], dlist[i][1], dlist[i][3]])
        elif (int(dlist[i][1].value) == int(day)) and (int(dlist[i][2].value)+100000 > int(time)+100000):
            data.append([dlist[i][0], dlist[i][1], dlist[i][3]])
        else:
            continue
    if len(data) < 1:
        print("no exist data to add")
        return 0
    return data


# engineController
def get_sync_time():
    try:
        ws = open_sheet("engine", "syncTime")
        return [i.value for i in list(ws.rows)[0][0:2]]
    except:
        print("can't read syncTime ")
        return -1


def get_type(mip):
    try:
        rs = open_sheet("engine", "types")
    except:
        print("can't read types")
        return -1
    for row in rs.rows:
        # empty rows 탐색 제외
        if row[0].value is None:
            break
        if row[0].value == mip:
            return row[1].value
    print("Invalid MIP")
    return -1


#barcode.xlsx 를 engine.xlsx에 동기화
def synchronization():
    st = get_sync_time()
    if st == -1:
        print("get sync time error")
        return -1
    #rawbarcode, date, groupid
    data = get_data_to_add(st[0], st[1])
    if data == 0 or data == -1:
        return -1
    try:
        w1b1 = open_file("engine")
        w2b1 = open_file("OwnedEngine")
    except:
        print("can't open .xlsx files")
        return -1
    try:
        w1s1 = w1b1["engineDB"]
        w1s2 = w1b1["engineGroup"]
        w1s3 = w1b1["syncTime"]
        w2s1 = w2b1["en"]
        glist = []
        for i in data:
            id = (i[0].value)[6:12]
            mip = (i[0].value)[12:]
            type = get_type(mip)
            day = i[1].value
            gid = str(i[2].value)
            if gid not in glist:
                glist.append(gid)
            # 엔진시리얼번호, mip, mip_type, 입고일, 포장일, 출고일, 출고설명, 그룹ID, 불량엔진bool타입, 비고
            #print([id, mip, type, day, day, "", "", gid, 0, ""])
            w1s1.append([id, mip, type, day, day, "", "", gid, 0, ""])
            # 엔진시리얼번호, mip, mip_type, 입고일, 포장일, 그룹ID, 불량엔진, 비고
            w2s1.append([id, mip, type, day, day, gid, 0, ""])
        for i in glist:
            #gio, location
            w1s2.append([i, ""])
        tm = datetime.datetime.now()
        #print(tm.strftime("%Y%m%d"), (tm.strftime("%X"))[0:2] + (tm.strftime("%X"))[3:5] + (tm.strftime("%X"))[6:8])
        w1s3.cell(row=1, column=1).value = tm.strftime("%Y%m%d")
        w1s3.cell(row=1, column=2).value = ((tm.strftime("%X"))[0:2] + (tm.strftime("%X"))[3:5] + (tm.strftime("%X"))[6:8])
        try:
            w1b1.save(get_path("engine"))
            w1b1.close()
            w2b1.save(get_path("OwnedEngine"))
            w2b1.close()
        except:
            print("save error")
            return -1
    except:
        print("can't write in sheets")
        return -1


def select_all_for_report(day):
    dl = defaultdict(list)
    rs = open_sheet("engine", "engineDB")
    for row in list(rs.rows)[1:]:
        # 아직 출고되지않았거나, 기간(a~b)에서 a날짜 이후의 데이터들에 대해 mip, 입고일, 출고일을 리스트에 append
        if (row[5].value == "") or (row[5].value is None) or (int(row[5].value) >= int(day)):
            dl[row[2].value].append([row[1].value, row[3].value, row[5].value])
        # print(row[5].value," ", int(day))
        # 역순 후 마지막행 제외
    return dl


def select_by_date(startdate, enddate):
    try:
        df = read_sheet("engine", "engineGroup")
    except:
        print("can't read engineGroup")
        return -1
    groups = defaultdict()
    for rname, row in df.iterrows():
        groups[row['groupID']] = row['Location']
    try:
        df = read_sheet("engine", "engineDB")
    except:
        print("can't open engineDB")
        return -1
    result = []
    for rname, row in df.iterrows():
        if (int(row['입고일']) >= int(startdate)) and (int(row['입고일']) <= int(enddate)):
            result.append([
                row['type'],
                row['mip'],
                row['eid'],
                row['입고일'],
                row['포장일'],
                row['출고일'],
                row['출고exp'],
                row['groupID'],
                groups[row['groupID']],
                "" if int(row['invalidEngine']) else "불량",
                row['비고']
            ])
    for i in range(len(result)):
        for j in range(11):
            if pd.isna(result[i][j]):
                result[i][j] = ""
    return result


# 보유 엔진에서 해당엔진삭제, 엔진데이터에서 출고일 수정 + 출고설명추가
def delete_row(en, comment, day):
    if (en is None) or (en == ""):
        print("empty en")
    wb1 = open_sheet("engine")
    ws1 = wb1['engineDB']
    wb2 = open_file("OwnedEngine")
    ws2 = wb2['en']
    # check for debug
    c1 = 0
    c2 = 0
    # eid 비교 후 같으면 해당 행 수정. 1부터 시작
    for r in range(1, ws1.max_row + 1):
        if str(ws1.cell(row=r, column=1).value) == en:
            ws1.cell(r, 6).value = day
            ws1.cell(r, 7).value = comment
            c1 = 1
            break
    # eid비교 후 같으면 해당 행 삭제.
    for r in range(1, ws2.max_row + 1):
        if str(ws2.cell(row=r, column=1).value) == en:
            ws2.delete_rows(r)
            c2 = 1
            break
    if (c1 * c2) == 0:
        print("not exist engine")
    wb1.save(get_path("engine"))
    wb1.close()
    wb2.save(get_path("OwnedEngine"))
    wb2.close()
    return 1


'''
def get_excellist():
    rs = open_sheet("engine", "engineDB")
    excelList = []
    tmpList = []
    for x in range(2, rs.max_row + 1):
        for y in range(1, rs.max_column + 1):
            cell = rs.cell(row=x, column=y).value
            if(y == 7):
                continue
            if cell is None:# or math.isnan(rs.cell(row=x, column=y).value):
                tmpList.append('')
            else:
                tmpList.append(cell)
        if tmpList[6] == '': #np.isnan
            tmpList.insert(7, '')
        else:
            #str = get_location(tmpList[6])
            if type(tmpList[6]) == type(int):
                str = get_location(tmpList[6])
            else:
                str = get_location(int(tmpList[6]))
            tmpList.insert(7, str)
        tmpList[0], tmpList[2] = tmpList[2], tmpList[0]
        excelList.append(tmpList)
        tmpList = []
    #excelList = excelList[1:]
    return excelList
'''

#기종, MIP, ENGINE, 입고일, 포장일, 출고일, 출고exp, GROUP, 위치, 불량엔진, 비고
def get_excellist():
    try:
        df = read_sheet("engine", "engineGroup")
    except:
        print("can't read engineGroup")
        return -1
    groups = defaultdict()
    for rname, row in df.iterrows():
        groups[row['groupID']] = row['Location']

    try:
        #df = pd.read_csv("./DB/engine.CSV", encoding='CP949')
        df = read_sheet("engine", "engineDB")
    except:
        print("can't read engineDB")
        return -1
    tmpList = []
    for rname, row in df.iterrows():
        tmpList.append([
            row['type'],
            row['mip'],
            row['eid'],
            row['입고일'],
            row['포장일'],
            row['출고일'],
            row['출고exp'],
            row['groupID'],
            groups[row['groupID']],
            "" if int(row['invalidEngine']) else "불량",
            row['비고']
        ])
    for i in range(len(tmpList)):
        for j in range(11):
            if pd.isna(tmpList[i][j]):
                tmpList[i][j] = ""
    return tmpList


def add_MIP(mip, types):
    wb1 = open_file("engine")
    _ = wb1.active
    ws1 = wb1["types"]
    ws1.append([mip,types])
    ws1.cell(ws1.max_row, 1).alignment = op.styles.Alignment(horizontal="center", vertical="center")
    ws1.cell(ws1.max_row, 2).alignment = op.styles.Alignment(horizontal="center", vertical="center")
    wb1.save(get_path("engine"))
    wb1.close()
    return


def set_invalid_engine(eng, exp):
    wb1 = open_file("engine")
    _ = wb1.active
    ws1 = wb1["engineDB"]

    #elist: 엔진 리스트
    #errorList: 입력으로 받은 값 중 잘못된 엔진 리스트
    elist = []
    errorList = []
    for row in ws1:
        elist.append(row[0].value)
    elist = elist[1:]
    for i in range(0, len(eng)):
        if not eng[i].isdigit():
            errorList.append(eng[i])
            continue
        int_value = int(eng[i])
        str_value = eng[i]
        if int_value in elist:
            idx = elist.index(int_value)
            ws1.cell(row=idx + 2, column=9, value=1)
            ws1.cell(row=idx + 2, column=10, value=exp[i])
        elif str_value in elist:
            idx = elist.index(str_value)
            ws1.cell(row=idx + 2, column=9, value=1)
            ws1.cell(row=idx + 2, column=10, value=exp[i])
        else:
            errorList.append(eng[i])
    wb1.save(get_path("engine"))
    wb1.close()
    return errorList

def get_today_date():
    dtNow = datetime.now()
    date = dtNow.date()
    date = str(date)
    date = date.replace('-', '')

    #test시 날짜 조작
    #date = int(date)
    #date -= 1
    #date = str(date)
    return date

#타입, mip로 딕셔너리 생성
def get_mip_dict():
    try:
        df = pd.read_excel("./DB/engine.xlsx", "types", header=None)
        #df = read_sheet("engine", "types")
        mipList = df[0].to_list()
        typeList = df[1].to_list()
        if len(mipList) != len(typeList):
            print("mip/type matching error")
    except:
        print("can't read engineGroup")
        return -1

    mipDict = {}
    count = len(mipList)
    for i in range(0, count):
        type = typeList[i]
        mip = mipList[i]
        if type in mipDict:
            if mip in mipDict[type]:
                #기종과 mip가 타입에 존재하는 경우
                continue
            else:
                mipDict[type][mip] = 0
        else:
            mipDict[type] = dict()
            mipDict[type][mip] = 0

    return mipDict

# 타입, MIP, 기초재고, 당기입고, 당기출고, 재고
def base_inventory_list(excelList, startdate, enddate):
    datetime_format = "%Y%m%d"
    sd = datetime.strptime(startdate, datetime_format)
    ed = datetime.strptime(enddate, datetime_format)
    tmpList = []
    receiveList = []
    releaseList = []

    for row in excelList:
        recv = datetime.strptime(row[3], datetime_format)
        if row[5] == '' or row[5] is None:
            release = ''
        else:
            release = datetime.strptime(row[5], datetime_format)

        # 시작일에 입고된 경우
        if recv == sd:
            # 시작일에 입고되고 불출되지 않은 경우: 기초재고 x / 당기입고
            if release == '' or release is None:
                receiveList.append(row)
                continue
            else:
                if release < sd:
                    print("리스트 오류")
                    continue
                elif release == sd:
                    # 시작일에 입고되고 당일 불출: 기초재고 X / 당기입고 및 출고
                    receiveList.append(row)
                    releaseList.append(row)
                    continue
                #release > sd
                if release == ed:
                    # 시작일 입고, 종료일 출고: 기초재고 X, 당기입고 및 출고
                    receiveList.append(row)
                    releaseList.append(row)
                    continue
                elif release > ed:
                    # 시작일에 입고되고 불출된 경우 되었는데 설정 기간보다 후에 불출된 경우: 기초재고 X / 당기입고
                    receiveList.append(row)
                    continue
                else:
                    # release < ed:
                    # 시작일에 입고되고, sd, ed 사이에 불출: 기초재고 X / 당기입고 및 출고
                    receiveList.append(row)
                    releaseList.append(row)
                    continue
            #있을 수 없는 케이스
            print("케이스 확인 요망")
            continue
        # 시작일 이전에 입고된 경우
        elif recv < sd:
            if release == '' or release is None:
                # 시작일이전에 입고되고 불출되지 않은 경우: 기초재고
                tmpList.append(row)
                continue
            else:
                if release < sd:
                    # 시작, 종료일 이전에 입출고 다 된 경우: 기초, 당기입출고 X
                    continue
                elif release == sd or release <= ed:
                    # 예전에 입고되어서 종료일 이전에 불출된 경우: 기초재고 및 당기 출고
                    tmpList.append(row)
                    releaseList.append(row)
                    continue
                if release > ed:
                    # 시작일 이전 입고 -> 마지막날 이후 출고: 기초재고
                    tmpList.append(row)
                    continue
            continue
        # 시작일 이후 입고
        elif recv > sd:
            if recv > ed:
                # 종료일 이후 입고 -> 기초재고 X
                continue
            else:
                # 시작일과 종료일 사이 입고
                if release == '' or release is None:
                    # ed, sd 사이에 입고되고 아직 출고되지 않음: 기초재고 X, 당기 입고
                    receiveList.append(row)
                    continue
                else:
                    if sd < release <= ed:
                        # 입,출고일이 sd, ed 사이, 출고일은 ed와 같을 수 있음: 기초재고 X, 당기입고 및 출고
                        receiveList.append(row)
                        releaseList.append(row)
                        continue
                    elif release > ed:
                        # 입,출고일이 sd, ed사이, ed이후에 출고: 기초재고 X, 당기입고
                        receiveList.append(row)
                        continue
            continue

    return tmpList, receiveList, releaseList

def inventory_payment(allList, sd, ed):
    # 재고수불 딕셔너리
    receiveDict = get_mip_dict()
    releaseDict = get_mip_dict()
    baseInv = get_mip_dict()
    #keyList
    first = []
    second = []
    first = list(baseInv.keys())
    # 재고수불 리스트
    baseList, receiveList, releaseList = base_inventory_list(allList, sd, ed)

    for tmp in first:
        second.append(list(baseInv[tmp].keys()))
    payment = []

    #기초재고 딕셔너리
    #invPayment: 기초재고 딕셔너리 {타입, {mip, 수량}}
    for row in baseList:
        tmpType = row[0]
        tmpMIP = row[1]
        if tmpType in baseInv:
            if tmpMIP in baseInv[tmpType]:
                baseInv[tmpType][tmpMIP] += 1
            else:
                print("해당 Type, MIP가 DB에 존재하지 않습니다.")
                return False, row[2]
        else:
            print("Type, MIP가 DB에 존재하지 않습니다.")
            return False + row[2]


    #mip별 딕셔너리 입출고 현황
    for row in allList:
        tmpType = row[0]
        tmpMIP = row[1]
        if row[3] == ed:
            if tmpType in receiveDict:
                if tmpMIP in receiveDict[tmpType]:
                    receiveDict[tmpType][tmpMIP] += 1
                else:
                    print("해당 Type, MIP가 DB에 존재하지 않습니다.")
                    return False, row[2]
            else:
                print("해당 Type, MIP가 DB에 존재하지 않습니다.")
                return False, row[2]

        if row[5] == ed:
            if tmpType in releaseDict:
                if tmpMIP in releaseDict[tmpType]:
                    releaseDict[tmpType][tmpMIP] += 1
                else:
                    print("해당 Type, MIP가 DB에 존재하지 않습니다.")
                    return False, row[2]
            else:
                print("해당 Type, MIP가 DB에 존재하지 않습니다.")
                return False, row[2]

    # 재고수불 리스트
    # 타입, mip, 기초재고, 당기입고, 당기출고, 재고
    for i in range(0, len(first)):
        for tmp in second[i]:
            tmpList = []
            sum = baseInv[first[i]][tmp] + receiveDict[first[i]][tmp] - releaseDict[first[i]][tmp]
            tmpList += [first[i], tmp, baseInv[first[i]][tmp], receiveDict[first[i]][tmp], releaseDict[first[i]][tmp], sum]
            payment.append(tmpList)

    return True, payment